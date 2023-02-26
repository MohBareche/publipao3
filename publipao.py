import tkinter as tk
from tkinter import *
from tkinter import ttk, font, messagebox as mb, filedialog as fd, simpledialog
from threading import Thread
import webbrowser
import ttkbootstrap as ttk
from ttkbootstrap.tableview import Tableview
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Querybox
from PIL import Image, ImageTk
import os
import re
import glob
import shutil
import win32com.client
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from pathlib import Path
from PyPDF2 import PdfMerger, PdfReader, PdfWriter

global data_name
global ent_list
global soum_list
global adj_list

to_list = []
from_list = []

window = ttk.Window(themename='darkly')
window_width = 1270
window_height = 730
screen_with = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
pos_x = (screen_with - window_width) // 2
pos_y = ((screen_height - window_height) // 2)-10
window.geometry(f"{window_width}x{window_height}+{pos_x}+{pos_y}")
window.title("Publipostage - Lettres de remerciement et lettre d'octroi")
window.resizable(0, 0)
window.iconbitmap('images/logo.ico')


def confirm_quitter():
    answer = mb.askyesno(title='Confirmation',
                         message='Êtes-vous sûr de vouloir quitter?')
    if answer:
        window.destroy()


def confirm_tout():
    mb.showinfo(title='Confirmation',
                message="Publipostage des lettres réalisé avec succès.")


def confirm_envoi_email():
    mb.showinfo(title='Confirmation', message="Courriel envoyé avec succès.")


def select_data_file():
    filetypes = (
        ('Fichier Excel', '*.xlsx'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Choisir une base de données',
        initialdir='./data',
        filetypes=filetypes
    )
    data_name = Path(filename)
    global wb
    wb = load_workbook(data_name)

    # liste chargés de projet
    ws_charg_proj = wb['Chargés de projet']
    list_charg_proj = []
    for cell in ws_charg_proj['B'][1:]:
        if cell.value != 'None':
            list_charg_proj.append(cell.value)
        lbl_message.grid(row=0, column=2)
        cmb_nom_charg_proj['values'] = list_charg_proj
        cmb_nom_charg_proj.configure(state='readonly')

    # liste gestionnaires
    ws_gest = wb['Gestionnaires']
    list_gestionnaires = []
    for cell in ws_gest['A'][1:]:
        if cell.value != 'None':
            list_gestionnaires.append(cell.value)
        cmb_nom_gestionnaire['values'] = list_gestionnaires
        cmb_nom_gestionnaire.current(0)
        cmb_nom_gestionnaire.configure(state='readonly')

    # liste secrétaires
    list_secret = []
    for cell in ws_gest['E'][1:]:
        if cell.value != 'None':
            list_secret.append(cell.value)
        cmb_secretaire['values'] = list_secret
        cmb_secretaire.current(0)
        cmb_secretaire.configure(state='readonly')

    select_remerc_file()
    select_octroi_file()

    lbl_message.configure(
        text='Base de données chargée avec succès...', width=55, relief='raised', bootstyle='inverse-success')


def load_data():
    new_window = Toplevel(window)
    new_window.title('Détail liste des entrepreneurs')
    new_window.iconbitmap('logo.ico')
    nom_charg_proj = cmb_nom_charg_proj.get()
    specialite = ""
    global sheet_data

    data = wb['Chargés de projet']
    for row in data.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == nom_charg_proj:
                specialite = data.cell(row=cell.row, column=5).value

    sheet_data = wb[specialite]

    r_set = [row for row in sheet_data.iter_rows(values_only=True)]

    l1 = r_set.pop(0)

    dv = ttk.tableview.Tableview(
        master=new_window,
        paginated=True,
        searchable=True,
        bootstyle='PRIMARY',
        pagesize=15,
        height=15,
        stripecolor=(colors.light, colors.dark)
    )
    # dv.grid(row=0, column=0, padx=5, pady=5)
    dv.pack(fill=tk.BOTH, expand='YES', padx=10, pady=10)
    dv.build_table_data(l1, r_set)
    dv.load_table_data()
    dv.autofit_columns()
    dv.autoalign_columns()


def moveTo(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())


def move_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def back_adj(from_list, to_list):
    selected_items = [from_list.get(idx) for idx in from_list.curselection()]
    for item in selected_items:
        to_list.insert(tk.END, item)
        from_list.delete(from_list.curselection())
        btn_adj_1.configure(state='normal')
        btn_adj_2.configure(state='disabled')


def soum_to_adj(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='normal')
    else:
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='normal')


def adj_to_soum(e):
    if not adj_list.get(0, tk.END):
        btn_adj_1.configure(state='disabled')
        btn_adj_2.configure(state='disabled')
    else:
        btn_adj_2.configure(state='normal')


def move_all(f_list, t_list):
    all_items = f_list.get(0, tk.END)
    f_list.delete(0, tk.END)
    for item in all_items:
        t_list.insert(tk.END, item)


def dbl_moveTo(e):
    ind_list = ent_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = ent_list.get(ind)
        ent_list.delete(ind)
        soum_list.insert(tk.END, val)


def dbl_moveBack(e):
    ind_list = soum_list.curselection()
    if ind_list:
        ind = ind_list[0]
        val = soum_list.get(ind)
        soum_list.delete(ind)
        ent_list.insert(tk.END, val)


def show_list_ent(e):
    ent_list.delete(0, tk.END)
    soum_list.delete(0, tk.END)
    adj_list.delete(0, tk.END)
    global nom_charg_proj
    global civ_charg_proj
    global courriel_charg_projet
    global tel_charg_proj
    global fonc_charg_proj
    global discipline

    nom_charg_proj = cmb_nom_charg_proj.get()
    data = wb['Chargés de projet']

    for row in data.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value == nom_charg_proj:

                civ_charg_proj = data.cell(row=cell.row, column=1).value
                courriel_charg_projet = data.cell(row=cell.row, column=3).value
                tel_charg_proj = data.cell(row=cell.row, column=4).value
                specialite = data.cell(row=cell.row, column=5).value
                fonc_charg_proj = data.cell(row=cell.row, column=6).value
                discipline = specialite

                if specialite == 'Voirie':
                    global list_ent_voirie
                    list_ent_voirie = wb['Voirie']
                    m_row = list_ent_voirie.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_voirie.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'Bâtiment':
                    global list_ent_bat
                    list_ent_bat = wb['Bâtiment']
                    m_row = list_ent_bat.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_bat.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

                if specialite == 'APA':
                    global list_ent_apa
                    list_ent_apa = wb['APA']
                    m_row = list_ent_apa.max_row
                    for i in range(2, m_row + 1):
                        nom_ent = list_ent_apa.cell(row=i, column=1)
                        ent_list.insert(tk.END, nom_ent.value)

    btn_load_detail_ent.configure(state='normal')


def select_remerc_file():
    global doc_remerc_name
    doc_remerc_name = 'Lettre_remerciement.docx'
    return doc_remerc_name


def select_octroi_file():
    global doc_octroi_name
    doc_octroi_name = 'Lettre_octroi.docx'
    return doc_octroi_name


def select_pv_ouverture_file():
    global doc_pv_ouvert_name
    filetypes = (
        ('Fichier PDF', '*.pdf'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title="Sélectionner le PV d'ouverture",
        initialdir='./pv',
        filetypes=filetypes
    )

    doc_pv_ouvert_name = Path(filename).name
    source = os.path.dirname(os.path.abspath(filename))
    destination = f"{os.getcwd()}\pv"

    if filename:
        lbl_pv_ouvert.configure(
            text="Procès verbal d'ouverture (OK)", bootstyle='SUCCESS')
        btn_gen_remerc.configure(state='normal')
        if lbl_pv_ca.cget("text") == "Procès verbal CA (OK)":
            btn_generer_tout.configure(state='normal')
        else:
            btn_generer_tout.configure(state='disabled')

        if source == destination:
            return
        else:
            shutil.move(os.path.join(source, doc_pv_ouvert_name), destination)
        return doc_pv_ouvert_name


def select_liste_commande_file():
    global doc_liste_commande_name
    filetypes = (
        ('Fichier PDF', '*.pdf'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title="Sélectionner la liste des commande",
        initialdir='./pv',
        filetypes=filetypes
    )
    doc_liste_commande_name = Path(filename).name
    source = os.path.dirname(os.path.abspath(filename))
    destination = f"{os.getcwd()}\pv"

    if filename:
        reader = PdfReader(filename)
        texte = reader.pages[0].extract_text()

        # Extraire le Numéro de contrat
        str_num_contrat = re.search(r"Numéro  : [\d]{1,}", texte).group()
        num_contrat = str_num_contrat.split(': ')[-1]
        txt_num_contrat.set(num_contrat)

        # Extraire le Numéro d'appel d'offres
        str_num_ao = re.search(
            r"Numéro de référence  : [\d]{1,}", texte).group()
        num_ao = str_num_ao.split(': ')[-1]
        txt_num_ao.set(num_ao)

        # # Extraire le titre du projet
        str_titre = re.search(
            r"Titre[\s\S]*?(?=(Important|Information))", texte).group()
        titre_retour = str_titre.split(': ')[-1]
        titre = titre_retour.replace("\n", " ")
        if titre.find('Montréal- Nord'):
            titre = titre.replace("Montréal- Nord", "Montréal-Nord")
        txt_titre_projet.set(titre)

        if source == destination:
            return
        else:
            shutil.move(os.path.join(
                source, doc_liste_commande_name), destination)
        return doc_liste_commande_name


def select_pv_ca_file():
    global doc_pv_ca_name
    filetypes = (
        ('Fichier Word', '*.doc'), ("All files", "*.*")
    )
    filename = fd.askopenfilename(
        title='Sélectionner le PV du CA',
        initialdir='./pv',
        filetypes=filetypes
    )

    doc_pv_ca_name = Path(filename).name
    source = os.path.dirname(os.path.abspath(filename))

    destination = f"{os.getcwd()}\pv"

    if filename:
        lbl_pv_ca.configure(text="Procès verbal CA (OK)", bootstyle='SUCCESS')
        btn_gen_octroi.configure(state='normal')
        if lbl_pv_ouvert.cget("text") == "Procès verbal d'ouverture (OK)":
            btn_generer_tout.configure(state='normal')
        else:
            btn_generer_tout.configure(state='disabled')

        if source == destination:
            return
        else:
            shutil.move(os.path.join(source, doc_pv_ca_name), destination)
        return doc_pv_ca_name


def get_secret_name(e):
    nom_redac = cmb_secretaire.get()
    return nom_redac


def initiales_gest(nom):
    init = ''.join(c for c in nom if c.isupper())
    return init


def initiales_secretaire(nom):
    cap = nom.split(' ')
    init = cap[0][0] + cap[1][0]
    return init.lower()


def erreur_msg():
    mb.showerror(title='Erreur',
                 message="Veuillez entrer les données manquantes.")


def gener_remerc():
    path = f'./gabarits/{doc_remerc_name}'
    doc = DocxTemplate(path)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

    ws_gestionnaires = wb['Gestionnaires']
    date = entry_cal.entry.get()
    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    nom_gest = cmb_nom_gestionnaire.get()
    init_gest = initiales_gest(nom_gest)
    global nom_redac
    global init_redac
    global titre_gest
    global fonction_gest
    global courriel_gestionnaire

    nom_redac = cmb_nom_charg_proj.get()
    init_redac = initiales_gest(nom_redac)

    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value
                courriel_gestionnaire = ws_gestionnaires.cell(
                    row=cell.row, column=4).value

    path_folder = './output/remerciement'
    isExist = os.path.exists(path_folder)
    if isExist:
        remerc_folder = shutil.rmtree('./output/remerciement')
    pathDOC = './output/remerciement/DOC'

    os.makedirs(pathDOC)

    for ent in list(soum_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_contrat,
            "nom_gestionnaire": nom_gest,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        nom_comp = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier = f"{num_contrat}_Lettre de remerciement - {nom_comp}.docx"

        doc.save(f'{pathDOC}/{nom_fichier}')

    path_doc_to_pdf_rem = f"{os.getcwd()}\output\\remerciement\DOC"
    path_rem = path_doc_to_pdf_rem
    word = win32com.client.Dispatch('Word.Application')
    for dirpath, dirnames, filenames in os.walk(path_rem):
        for file in filenames:
            if file.lower().endswith(".docx"):
                new_name = file.replace(".docx", ".pdf")
                in_file = (dirpath + '/' + file)
                new_file = (dirpath + '/' + new_name)
                doc = word.Documents.Open(in_file)
                doc.SaveAs(new_file, FileFormat=17)
                doc.Close()
            if file.lower().endswith(".doc"):
                new_name = file.replace(".doc", ".pdf")
                in_file = (dirpath + '/' + file)
                new_file = (dirpath + '/' + new_name)
                doc = word.Documents.Open(in_file)
                doc.SaveAs(new_file, FileFormat=17)
                doc.Close()
    word.Quit()

    source_dir = path_doc_to_pdf_rem
    target_dir = os.getcwd()

    for fname in os.listdir(source_dir):
        if fname.lower().endswith('.pdf'):
            shutil.move(os.path.join(source_dir, fname), target_dir)

    pv_ouverture = f"./pv/{doc_pv_ouvert_name}"
    pdf_pv_ouverture = open(pv_ouverture, 'rb')

    pdfs = glob.glob('*.pdf')

    for pdf in pdfs:
        merger = PdfMerger()
        merger.append(pdf)
        merger.append(pdf_pv_ouverture)
        name = pdf.split(".")[0]
        merger.write(f"{name}_.pdf")
        merger.close()

    pdf_folder = 'PDF'
    remerc_folder = './output/remerciement'
    pathPDF = os.path.join(remerc_folder, pdf_folder)
    os.makedirs(pathPDF)

    for f in glob.glob('./*_.pdf'):
        shutil.move(f, pathPDF)

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)


def gener_octroi():
    path_gabarit = f'./gabarits/{doc_octroi_name}'
    doc = DocxTemplate(path_gabarit)
    compagnies = {}
    ws = wb[discipline]
    for row in ws.iter_rows(min_row=2, values_only=True):
        company_name = row[0]
        company_data = {
            "nom_de_compagnie": row[0],
            "adresse": row[1],
            "ville": row[2],
            "code_postal": row[3],
            "courriel": row[4],
            "representant": row[5],
            "civilite": row[6],
            "fonction": row[7]
        }
        compagnies[company_name] = company_data

        ws_gestionnaires = wb['Gestionnaires']

    pv_ca = f"./pv/{doc_pv_ca_name}"
    shutil.move(pv_ca, './')

    filename = doc_pv_ca_name
    filenamePDF = filename.split('.')[0]
    path = os.getcwd()
    in_file = f"{path}\{filename}"
    out_file = f"{path}\{filenamePDF}"
    pv_ca_pdf = f"{filenamePDF}.pdf"

    wdFormatPDF = 17
    word = win32com.client.Dispatch('Word.Application')
    doc_doc = word.Documents.Open(in_file)
    doc_doc.SaveAs(out_file, FileFormat=wdFormatPDF)
    doc_doc.Close()
    word.Quit()
    shutil.move(in_file, './pv')

    reader = PdfReader(f"{out_file}.pdf")
    texte = reader.pages[0].extract_text()
    resolution = re.search(r"CA[\d]{2}\s[\d]{2}\s[\d]{2,4}", texte).group()
    date_resolution = re.search(
        r"[\d]{1,2}\s(?:janvier|février|mars|avril|mai|juin|juillet|août|septembre|octobre|novembre|décembre)\s[\d]{4}", texte).group()

    date = entry_cal.entry.get()
    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    num_ao = entry_num_ao.get()
    nom_gest = cmb_nom_gestionnaire.get()
    nom_charg_proj = cmb_nom_charg_proj.get()
    init_gest = initiales_gest(nom_gest)
    autre_personne = txt_cc_lettre_octroi.get(1.0, "end-1c")

    ws_gestionnaires = wb['Gestionnaires']
    for row in ws_gestionnaires.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == nom_gest:
                global titre_gest
                global fonction_gest
                titre_gest = ws_gestionnaires.cell(
                    row=cell.row, column=2).value
                fonction_gest = ws_gestionnaires.cell(
                    row=cell.row, column=3).value

    path_folder = './output/octroi'
    isExist = os.path.exists(path_folder)
    if isExist:
        octroi_folder = shutil.rmtree('./output/octroi')
    pathDOC = './output/octroi/DOC'
    os.makedirs(pathDOC)

    for ent in list(adj_list.get(0, tk.END)):
        doc.render({
            "date": date,
            "titre": titre_projet,
            "num_contrat": num_contrat,
            "num_ao": num_ao,
            "nom_gestionnaire": nom_gest,
            "autre_personne": autre_personne,
            "titre_gest": titre_gest,
            "fonction_gest": fonction_gest,
            "init_gest": init_gest,
            "init_redac": init_redac,
            "resolution": resolution,
            "date_resolution": date_resolution,
            "civ_charg_proj": civ_charg_proj,
            "nom_charg_proj": nom_charg_proj,
            "tel_charg_proj": tel_charg_proj,
            "fonc_charg_proj": fonc_charg_proj,
            "civilite": compagnies[ent]['civilite'],
            "representant": compagnies[ent]['representant'],
            "nom_de_compagnie": compagnies[ent]['nom_de_compagnie'],
            "adresse": compagnies[ent]['adresse'],
            "ville": compagnies[ent]['ville'],
            "code_postal": compagnies[ent]['code_postal'],
            "courriel": compagnies[ent]['courriel']
        })
        global nom_comp_adj
        global nom_lettre_pdf
        global nom_final
        nom_comp_adj = f'{compagnies[ent]["nom_de_compagnie"]}'
        nom_fichier_doc = f"{num_contrat}_Lettre d'adjudication - {nom_comp_adj}.docx"
        doc.save(f'{pathDOC}/{nom_fichier_doc}')
        nom_lettre_pdf = f"{num_contrat}_Lettre d'adjudication - {nom_comp_adj}.pdf"
        nom_final = f"{num_contrat}_Lettre d'adjudication - {nom_comp_adj}"

    path_doc_to_pdf_oct = f"{os.getcwd()}\output\octroi\DOC"
    path_oct = path_doc_to_pdf_oct

    word = win32com.client.Dispatch('Word.Application')
    for dirpath, dirnames, filenames in os.walk(path_oct):
        for file in filenames:
            if file.lower().endswith(".docx"):
                new_name = file.replace(".docx", ".pdf")
                in_file = (dirpath + '/' + file)
                new_file = (dirpath + '/' + new_name)
                doc = word.Documents.Open(in_file)
                doc.SaveAs(new_file, FileFormat=17)
                doc.Close()
            if file.lower().endswith(".doc"):
                new_name = file.replace(".doc", ".pdf")
                in_file = (dirpath + '/' + file)
                new_file = (dirpath + '/' + new_name)
                doc = word.Documents.Open(in_file)
                doc.SaveAs(new_file, FileFormat=17)
                doc.Close()
    word.Quit()

    source_dir = path_doc_to_pdf_oct
    target_dir = os.getcwd()

    for fname in os.listdir(source_dir):
        if fname.lower().endswith('.pdf'):
            shutil.move(os.path.join(source_dir, fname), target_dir)

    pdfs = [nom_lettre_pdf, pv_ca_pdf]
    merger = PdfWriter()

    for pdf in pdfs:
        merger.append(pdf)

    merger.write(f"{nom_final}_.pdf")
    merger.close()

    pdf_folder = 'PDF'
    octroi_folder = './output/octroi'
    pathPDF = os.path.join(octroi_folder, pdf_folder)
    os.makedirs(pathPDF)

    for f in glob.glob('./*_.pdf'):
        shutil.move(f, pathPDF)

    for f in os.listdir('./'):
        if f.endswith('.pdf'):
            os.remove(f)


def gener_tout():
    gener_remerc()
    gener_octroi()


def envoi_courriel():
    global password
    path_key = './key.txt'
    isExist = os.path.isfile(path_key)
    if isExist:
        with open(path_key, 'r') as file:
            password = file.read().rstrip()
            if len(password) == 16 and password.isalpha():
                envoyer()
            else:
                mb.showerror(
                    title='Erreur', message='Votre mot de passe est incorrect. Corrigez-le et reessayez !')
                envoi_courriel()
    else:
        new_window = tk.Tk()
        new_window.iconbitmap('logo.ico')
        new_window.withdraw()
        password = simpledialog.askstring(
            title="Mot de passe", prompt="Entrez votre mot de passe", parent=new_window)
        new_window.destroy()
        if (password == None) or (len(password) != 16) or not password.isalpha():
            mb.showerror(
                title='Erreur', message='Vous devez entrer un mot de passe valide pour pouvoir envoyer un courriel.')
            return
        else:
            envoyer()


def envoyer():
    num_contrat = entry_num_contrat.get()
    nom_gestionnaire = cmb_nom_gestionnaire.get()
    nom_secretaire = cmb_secretaire.get()
    prenom_gestionnaire = nom_gestionnaire.split(' ')[0]

    global courriel_secretaire
    data = wb['Gestionnaires']

    for row in data.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if cell.value == nom_secretaire:
                courriel_secretaire = data.cell(row=cell.row, column=6).value

    fromaddr = courriel_charg_projet
    toaddr = [courriel_gestionnaire, courriel_secretaire]
    redact = nom_charg_proj.split(' ')[0]

    dir_path_rem = "./output/remerciement/PDF"
    dir_path_oct = "./output/octroi/PDF"

    files_rem = []
    for fname in os.listdir(dir_path_rem):
        if fname.lower().endswith('.pdf'):
            files_rem.append(fname)

    files_oct = []
    for fname in os.listdir(dir_path_oct):
        if fname.lower().endswith('.pdf'):
            files_oct.append(fname)

    msg = MIMEMultipart()
    msg['To'] = fromaddr
    msg['From'] = ", ".join(toaddr)
    msg['Subject'] = f"{num_contrat} - Lettres de remerciement et lettre d'octroi pour signature"

    text = f"""
Bonjour {prenom_gestionnaire},

Vous trouverez en pièces jointes des documents pour signature.

Si vous avez des questions, n'hésitez pas à communiquer avec moi.

Cordialement.

{redact}.
"""
    body = MIMEText(text, 'plain')
    msg.attach(body)

    for f in files_rem:  # ajout des fichiers attachés au message
        file_path = os.path.join(dir_path_rem, f)
        attachment = MIMEApplication(
            open(file_path, "rb").read(), _subtype="pdf")
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=f)
        msg.attach(attachment)

    for f in files_oct:  # ajout des fichiers attachés au message
        file_path = os.path.join(dir_path_oct, f)
        attachment = MIMEApplication(
            open(file_path, "rb").read(), _subtype="pdf")
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', 'attachment', filename=f)
        msg.attach(attachment)

    # creates SMTP session
    s = smtplib.SMTP('smtp.gmail.com', 587)
    # start TLS for security
    s.starttls()
    # Authentification
    s.login(fromaddr, password)
    # Convert. Multipart msg en chaine de caractère
    text = msg.as_string()
    # envoi mail
    s.sendmail(fromaddr, toaddr, text)
    # terminer la session
    s.quit()


def enable_btn_disabled():
    btn_open_folder.configure(state='normal')
    btn_env_courriel.configure(state='normal')


def update_theme(e):
    window.style.theme_use(nom_theme.get())


def open_folder():
    path = './output'
    os.system(f'start {os.path.realpath(path)}')


def show_and_run(func):
    func()


def run_function(func, btn):
    progressbar = ttk.Progressbar(frame_progress, orient='horizontal',
                                  mode='indeterminate', length=280, bootstyle="info-striped")
    progressbar.grid(row=0, column=0, padx=10, pady=10)
    progressbar.start(interval=10)
    frame_progress.configure(text='Publipostage en cours. Patientez . . .')
    show_and_run(func)
    enable_btn_disabled()
    progressbar.destroy()
    frame_progress.configure(text='Statut')
    confirm_tout()


def run_email(func, btn):
    progressbar = ttk.Progressbar(frame_progress, orient='horizontal', mode='indeterminate', length=280,bootstyle="success-striped")
    progressbar.grid(row=0, column=0, padx=10, pady=10)
    progressbar.start(interval=10)
    show_and_run(func)
    frame_progress.configure(text='Envoi en cours. Patienter . . .')
    progressbar.destroy()
    frame_progress.configure(text='Statut')
    confirm_envoi_email()


def generer(func, btn):
    global titre_projet
    global num_contrat
    global num_ao

    titre_projet = entry_titre_projet.get()
    num_contrat = entry_num_contrat.get()
    num_ao = entry_num_ao.get()
    nom_charg_proj = cmb_nom_charg_proj.get()

    if (titre_projet and num_contrat and num_ao and nom_charg_proj):
        Thread(target=run_function, args=(func, btn)).start()
    else:
        erreur_msg()


def envoi(func, btn):
    answer = mb.askyesno(title='Confirmation', message='Êtes-vous sûr de vouloir envoyer le courriel ?')
    if answer:
        Thread(target=run_email, args=(func, btn)).start()
    else: 
        return


def reinit():
    lbl_message.configure(text="Choisir d'abord la base de données pour commencer.",
                          width=50, relief='raised', bootstyle="inverse-danger")
    entry_titre_projet.delete(0, END)
    entry_titre_projet.focus_set()
    entry_num_contrat.delete(0, END)
    entry_num_ao.delete(0, END)
    ent_list.delete(0, END)
    soum_list.delete(0, END)
    adj_list.delete(0, END)
    cmb_nom_charg_proj.set('')
    cmb_nom_gestionnaire.set('')
    cmb_secretaire.set('')
    cmb_nom_charg_proj['values'] = []
    cmb_nom_gestionnaire['values'] = []
    cmb_secretaire['values'] = []
    lbl_pv_ouvert.configure(
        text="Procès verbal d'ouverture (.pdf)", bootstyle='default')
    lbl_pv_ca.configure(text='Procès verbal CA (.doc)', bootstyle='default')
    btn_gen_remerc.configure(state='disabled')
    btn_gen_octroi.configure(state='disabled')
    btn_generer_tout.configure(state='disabled')
    btn_env_courriel.configure(state='disabled')
    btn_open_folder.configure(state='disabled')
    btn_load_detail_ent.configure(state='disabled')


def choix():
    pop.destroy()


def apropos():
    global pop
    pop = Toplevel(window)
    pop.overrideredirect(1)
    pop_width = 400
    pop_height = 150
    scr_with = pop.winfo_screenwidth()
    scr_height = pop.winfo_screenheight()
    pos_x = (scr_with - pop_width) // 2
    pos_y = ((scr_height - pop_height) // 2)-10
    pop.geometry(f"{pop_width}x{pop_height}+{pos_x}+{pos_y}")
    pop.title("À Propos")
    pop.resizable(0, 0)
    pop.iconbitmap('images/logo.ico')
    pop.grid_rowconfigure(0, weight=1)
    pop.grid_columnconfigure(0, weight=1)
    frame = ttk.Frame(pop)
    frame.grid(row=0, column=0, sticky='NEWS', padx=20, pady=10)
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)
    copyright = u"\u00A9"
    texte = f"Mohamed Bareche {copyright} 2023"

    label1 = Label(
        frame, text="Application conçue et réalisée par", bg='#59CE8F')
    label1.grid(row=0, column=0, sticky='NEWS')
    label2 = Label(frame, text=texte, bg='white')
    label2.grid(row=1, column=0, sticky='NEWS')

    img = Image.open("images/logo.png")
    image1 = img.resize((70, 70), Image.LANCZOS)
    test = ImageTk.PhotoImage(image1)

    label_img = ttk.Label(frame, image=test)
    label_img.image = test
    label_img.grid(row=0, column=1, rowspan=3, padx=10, sticky='NEWS')
    buttonOk = ttk.Button(frame, text="Ok", command=choix, bootstyle='PRIMARY')
    buttonOk.grid(row=2, column=0, padx=10, pady=10, sticky='NEWS')
    pop.configure(bg='#59CE8F')


def open_webpage():
    webbrowser.open('https://www.seao.ca')


# default font
window.defaultFont = font.nametofont("TkDefaultFont")
window.defaultFont.configure(family="Arial", size=11)

colors = window.style.colors
themes = window.style.theme_names()
nom_theme = ttk.StringVar(value=window.style.theme_use())

# ********************************************************************************************************************

frame_first = ttk.Frame(window, width=750,  height=100)
frame_first.grid(row=0, column=0, columnspan=2, sticky='NEWS', padx=20, pady=7)

frame_data = ttk.LabelFrame(
    frame_first, text='Choix de la base de données', width=700)
frame_data.grid(row=0, column=0, sticky='W', padx=5, pady=7)

lbl_load_data = ttk.Label(frame_data, text='Choisir un fichier')
lbl_load_data.grid(row=0, column=0)
btn_load_data = ttk.Button(
    frame_data, text='Sélectionner...', command=select_data_file, bootstyle='PRIMARY, OUTLINE')
btn_load_data.grid(row=0, column=1)

lbl_message = ttk.Label(frame_data,
                        text="Choisir d'abord la base de données pour commencer.", width=55, relief='raised', bootstyle="inverse-danger")
lbl_message.grid(row=0, column=2)

frame_progress = ttk.LabelFrame(
    frame_first, width=300, height=50, text='Statut')
frame_progress.grid(row=0, column=1, padx=20, pady=7, sticky='NEWS')

frame_theme = ttk.LabelFrame(
    frame_first, text='Sélectionner un thème', width=300)
frame_theme.grid(row=0, column=2,  sticky='E', padx=0, pady=7)

for theme in themes:
    cmb_theme = ttk.Combobox(
        frame_theme, state='readonly', textvariable=nom_theme, values=themes)
    cmb_theme.current(11)
    cmb_theme.grid(row=0, column=0, padx=7, pady=10)
    cmb_theme.bind('<<ComboboxSelected>>', update_theme)

for widget in frame_data.winfo_children():
    widget.grid_configure(padx=5, pady=7)

for widget in frame_progress.winfo_children():
    widget.grid_configure(padx=5, pady=7)

for widget in frame_theme.winfo_children():
    widget.grid_configure(padx=5, pady=7)

# ********************************************************************************************************************

frame_info_projet = ttk.LabelFrame(
    window,  text='Informations sur le projet', height=100)
frame_info_projet.grid(row=1, column=0, columnspan=2,
                       sticky='NEWS', padx=20, pady=7)

lbl_load_info_proj = ttk.Label(frame_info_projet, text='Charger info.')
lbl_load_info_proj.grid(row=0, column=0)
btn_load_info_proj = ttk.Button(
    frame_info_projet, text='Sélectionner...', command=select_liste_commande_file, bootstyle='PRIMARY, OUTLINE')
btn_load_info_proj.grid(row=0, column=1)

txt_titre_projet = tk.StringVar()
lbl_titre_projet = ttk.Label(frame_info_projet, text='Titre projet')
lbl_titre_projet.grid(row=0, column=2)
entry_titre_projet = ttk.Entry(
    frame_info_projet, textvariable=txt_titre_projet, width=55)
entry_titre_projet.grid(row=0, column=3)

txt_num_contrat = tk.StringVar()
lbl_num_contrat = ttk.Label(frame_info_projet, text='No contrat')
lbl_num_contrat.grid(row=0, column=4)
entry_num_contrat = ttk.Entry(frame_info_projet, textvariable=txt_num_contrat)
entry_num_contrat.grid(row=0, column=5)

txt_num_ao = tk.StringVar()
lbl_num_ao = ttk.Label(frame_info_projet, text="No A.O.")
lbl_num_ao.grid(row=0, column=6)
entry_num_ao = ttk.Entry(frame_info_projet, textvariable=txt_num_ao)
entry_num_ao.grid(row=0, column=7)

photo = ImageTk.PhotoImage(Image.open(
    "images/www.png").resize((30, 30), Image.LANCZOS))
btn_open_webpage = tk.Button(
    frame_info_projet, image=photo, command=open_webpage, borderwidth=0)
btn_open_webpage.grid(row=0, column=8, rowspan=2, sticky='W')

for widget in frame_info_projet.winfo_children():
    widget.grid_configure(padx=10, pady=10)

# ********************************************************************************************************************

frame_info_charg_proj_sign_date = ttk.LabelFrame(
    window, text='Informations diverses [ Chargé de projet, signataire et date de rédaction ]', height=100)
frame_info_charg_proj_sign_date.grid(
    row=2, column=0, columnspan=2, sticky='NEWS', padx=20, pady=7)

lbl_nom_charg_proj = ttk.Label(
    frame_info_charg_proj_sign_date, text='Chargé(e) de projet')
lbl_nom_charg_proj.grid(row=0, column=0, sticky='s')
cmb_nom_charg_proj = ttk.Combobox(
    frame_info_charg_proj_sign_date, width=25, bootstyle='PRIMARY')
cmb_nom_charg_proj.grid(row=1, column=0, sticky='n')
cmb_nom_charg_proj.bind("<<ComboboxSelected>>", show_list_ent)

# informations sur le signataire (gestionnaire)
lbl_nom_gestionnaire = ttk.Label(
    frame_info_charg_proj_sign_date, text='Signataire (Gestionnaire)')
lbl_nom_gestionnaire.grid(row=0, column=1, sticky='s')
cmb_nom_gestionnaire = ttk.Combobox(frame_info_charg_proj_sign_date, width=25)
cmb_nom_gestionnaire.grid(row=1, column=1, sticky='n')

# date de rédaction
lbl_date = ttk.Label(frame_info_charg_proj_sign_date,
                     text="Date de rédaction")
lbl_date.grid(row=0, column=2, sticky='s')

entry_cal = ttk.DateEntry(
    frame_info_charg_proj_sign_date, dateformat='%d %B %Y')
entry_cal.grid(row=1, column=2, sticky='n')

# Secrétaire en CC dans le courriel
lbl_redac = ttk.Label(frame_info_charg_proj_sign_date,
                      text='Secrétaire en CC dans courriel')
lbl_redac.grid(row=0, column=3, sticky='s')
cmb_secretaire = ttk.Combobox(
    frame_info_charg_proj_sign_date, width=25, bootstyle='PRIMARY')
cmb_secretaire.grid(row=1, column=3, sticky='n')

# Personnes en CC dans la lettre d'octroi
lbl_cc_lettre_octroi = ttk.Label(frame_info_charg_proj_sign_date,
                                 text="Pers. en CC / lettre d'octroi excluant chargé de projet")
lbl_cc_lettre_octroi.grid(row=0, column=4, sticky='s')
txt_cc_lettre_octroi = tk.Text(
    frame_info_charg_proj_sign_date, height=3, width=30)
txt_cc_lettre_octroi.grid(row=1, column=4, sticky='n')

for widget in frame_info_charg_proj_sign_date.winfo_children():
    widget.grid_configure(padx=15, pady=10)

# ********************************************************************************************************************

frame_soumission = ttk.LabelFrame(
    window, text='Informations sur les soumissionnaires', height=200)
frame_soumission.grid(row=3, column=0, columnspan=2,
                      sticky='NEWS', padx=20, pady=7)

lbl_list_ent = ttk.Label(frame_soumission, text='Liste des entrepreneurs')
lbl_list_ent.grid(row=0, column=0)
ent_list = tk.Listbox(frame_soumission,
                      width=40, font=('Arial', 10))
ent_list.grid(row=1, column=0)

frame_group_btn1 = ttk.Frame(frame_soumission)
frame_group_btn1.grid(row=1, column=1, rowspan=4)

btn_load_detail_ent = ttk.Button(frame_group_btn1, text='Détails',
                                 bootstyle='SUCCESS', width=7, state='disabled', command=load_data)
btn_load_detail_ent.grid(row=0, column=0)

btn_1 = ttk.Button(frame_group_btn1, text='>', bootstyle='DANGER, OUTLINE',
                   width=7, command=lambda: moveTo(ent_list, soum_list))
btn_1.grid(row=1, column=0, pady=5)

btn_2 = ttk.Button(frame_group_btn1, text='>>', bootstyle='DANGER, OUTLINE',
                   width=7, command=lambda: move_all(ent_list, soum_list))
btn_2.grid(row=2, column=0, pady=5)

btn_3 = ttk.Button(frame_group_btn1, text='<', bootstyle='DANGER, OUTLINE',
                   width=7, command=lambda: moveTo(soum_list, ent_list))
btn_3.grid(row=3, column=0, pady=5)

btn_4 = ttk.Button(frame_group_btn1, text='<<', bootstyle='DANGER, OUTLINE',
                   width=7, command=lambda: move_all(soum_list, ent_list))
btn_4.grid(row=4, column=0, pady=5)

lbl_list_soum = ttk.Label(frame_soumission, text='Liste des soumissionnaires')
lbl_list_soum.grid(row=0, column=2)
soum_list = tk.Listbox(frame_soumission, width=40, font=('Arial', 10))
soum_list.grid(row=1, column=2)

frame_group_btn2 = ttk.Frame(frame_soumission)
frame_group_btn2.grid(row=1, column=3)

btn_adj_1 = ttk.Button(frame_group_btn2, text='Octroyer>', state='disabled', bootstyle='SUCCESS, OUTLINE',
                       width=10, command=lambda: [move_adj(soum_list, adj_list), soum_to_adj])
btn_adj_1.grid(row=0, column=0, pady=5)

btn_adj_2 = ttk.Button(frame_group_btn2, text='<Retirer', state='disabled', bootstyle='SUCCESS, OUTLINE',
                       width=10, command=lambda: [back_adj(adj_list, soum_list), adj_to_soum])
btn_adj_2.grid(row=1, column=0, pady=5)

lbl_adj = ttk.Label(frame_soumission, text='Entreprise adjugée')
lbl_adj.grid(row=0, column=4)
adj_list = tk.Listbox(frame_soumission, width=40, font=('Arial', 10))

for widget in frame_soumission.winfo_children():
    widget.grid_configure(padx=10, pady=0)

adj_list.grid(row=1, column=4, pady=10)

# ********************************************************************************************************************

frame_remerc_octroi = ttk.LabelFrame(
    window, text='Lettres de remerciement et octroi',  height=200)
frame_remerc_octroi.grid(row=4, column=0, sticky='NEWS', padx=20, pady=7)

frame_remerc = ttk.Frame(
    frame_remerc_octroi)
frame_remerc.grid(row=0, column=0, sticky='N')

# PV Ouverture Remerciements
lbl_pv_ouvert = ttk.Label(
    frame_remerc, text="Procès verbal d'ouverture (.pdf)", width=30)
lbl_pv_ouvert.grid(row=0, column=0, stick='W')

btn_pv_ouvert = ttk.Button(
    frame_remerc, text='Sélectionner...', bootstyle='PRIMARY, OUTLINE', command=select_pv_ouverture_file, width=15)
btn_pv_ouvert.grid(row=0, column=1, sticky='W')

btn_gen_remerc = ttk.Button(
    frame_remerc, text='Générer les lettres de remerciement', bootstyle='SUCCESS', width=35, state='disabled',
    command=lambda: generer(gener_remerc, btn_gen_remerc))
btn_gen_remerc.grid(row=0, column=2, sticky='E')

for widget in frame_remerc.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# --------------------------------------------------------------------------------------------------

frame_octroi = ttk.Frame(
    frame_remerc_octroi)
frame_octroi.grid(row=1, column=0, sticky='S')
# PV CA Octroi
lbl_pv_ca = ttk.Label(frame_octroi, text='Procès verbal CA (.doc)', width=30)
lbl_pv_ca.grid(row=0, column=0, sticky='W')
btn_pv_ca = ttk.Button(
    frame_octroi, text='Sélectionner...', bootstyle='PRIMARY, OUTLINE', command=select_pv_ca_file, width=15)
btn_pv_ca.grid(row=0, column=1, sticky='W')

btn_gen_octroi = ttk.Button(
    frame_octroi, text="Générer la lettre d'octroi", bootstyle='SUCCESS', width=35, state='disabled',
    command=lambda: generer(gener_octroi, btn_gen_octroi))
btn_gen_octroi.grid(row=0, column=2, sticky="E")

for widget in frame_octroi.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# ********************************************************************************************************************

frame_btns = ttk.Frame(window, height=100, width=300)
frame_btns.grid(row=4, column=1, padx=20, pady=7, sticky='E')

btn_generer_tout = ttk.Button(
    frame_btns, text='Générer tout (remerciement et octroi)', bootstyle='PRIMARY',
    command=lambda: generer(gener_tout, btn_generer_tout), state='disabled')
btn_generer_tout.grid(row=0, column=0, columnspan=3, sticky='NEWS')

btn_env_courriel = ttk.Button(frame_btns, text='Envoyer courriel', state='disabled',
                              bootstyle='INFO', command=lambda: envoi(envoi_courriel, btn_env_courriel))
btn_env_courriel.grid(row=1, column=0, sticky='NEWS')

btn_open_folder = ttk.Button(frame_btns, text='Explorer', state='disabled',
                             bootstyle='INFO', command=open_folder)
btn_open_folder.grid(row=1, column=1, columnspan=2, sticky='NEWS')

btn_reinit = ttk.Button(frame_btns, text='Réinitialiser',
                        bootstyle='WARNING', command=reinit)
btn_reinit.grid(row=2, column=0, sticky='NEWS')

btn_quitter = ttk.Button(frame_btns, text='Quitter',
                         bootstyle='DANGER', command=confirm_quitter)
btn_quitter.grid(row=2, column=1,  sticky='NEWS')

btn_apropos = ttk.Button(frame_btns, text='?',
                         bootstyle='PRIMARY', command=apropos)
btn_apropos.grid(row=2, column=2,  sticky='NEWS')

for widget in frame_btns.winfo_children():
    widget.grid_configure(padx=5, pady=5)

ent_list.bind('<Double-Button>', dbl_moveTo)
soum_list.bind('<Double-Button>', dbl_moveBack)
soum_list.bind('<<ListboxSelect>>', soum_to_adj)
adj_list.bind('<<ListboxSelect>>', adj_to_soum)
cmb_secretaire.bind('<<ComboboxSelected>>', get_secret_name)

window.mainloop()